# Importamos las librerias que utilizaremos durante el programa
import msvcrt
import math
import os
from openpyxl import load_workbook
from openpyxl import Workbook
import random
import pandas as pd
import time
# Importamos los TAD
from arboles import Arbol
from persona import solicitudCliente
from persona import operario


pino = Arbol('Pino', 0.78, 3.96, 2042.03)
algarrobo = Arbol('Algarrobo', 1.76, 8.83, 6847.69)
eucalipto = Arbol('Eucalipto', 1.76, 8.83, 7333.65)
Urunday = Arbol('Urunday', 0.38, 1.92, 2116.64)
quebracho = Arbol('Quebracho Colorado', 0.19, 0.98, 1197.73)


def PosCamion(Tchip):
    # datos Vi= 5 M/s, Yi=3,5 M, 60° grados de inclinacion. Altura del camion 4,3 metros largo del camion 2,60 metros
    # Formulas El eje y (Yf =Yi + VYi *T + 12*g*T2) (VyF =Vyi +g*T)  (Vyi= sen ∝*Vi) eje x (Xf =Xi + VXi *T ) (VxF= cos ∝* Vi)
    if Tchip >= 8:
        Vi = 10
    elif Tchip <= 6 and Tchip >= 5:
        Vi = 8
    elif Tchip <= 4:
        Vi = 6
    Acamion = 4.3
    Lcamino = 2.6
    Viy = Vi*(math.sin(60*math.pi/180))  # Velocidad en el eje y
    t1 = Viy/9.8  # Tiempo en segundos que tarda en llegar al punto mas alto
    Ymax = (3.5+(Viy*t1))-(4.9*(t1)**2)  # Ymax en metros
    Vx = Vi*(math.cos(60*math.pi/180))  # Velocidad en el eje x
    # Tiempo en segundos que tarda en llegar a la altura maxima del camion
    t2 = math.sqrt((Ymax-Acamion)/4.9)
    tt = t1+t2  # Tiempo total en segundos que tardaran en caer al acoplado
    Xmax = -(Lcamino/2)+(Vx*tt)  # Xmax en metros
    print("El camion se debe ubicar a ", round(Xmax, 2),
          " metros de distancia de la maquina trituradora")


# Funcion de bienvenida al usuario operador
def ingresarCredenciales():
    operario1 = operario.ingresarDatos()
    operario1.mostrarDatos()
    return operario1

# Crea el archivo execel para los pedidos de los clientes


def crear_excel_pedidos():
    wb = Workbook()
    ws = wb.active
    ws.title = "Pedidos-Clientes"
    ws.append(["Fecha", "ID Cliente", "Nombre Cliente", "Apellido Cliente",
              "Cantidad de Chips(KG)", "Tamaño Chips", "Tipo Madera"])
    wb.save("Pedidos-Clientes.xlsx")


# Crea el archivo execel para los pedidos de las diferentes producciones
def crear_excel_prodcciones():
    wb = Workbook()
    ws = wb.active
    ws.title = "Historial-Producciones"
    ws.append(["Fecha", "ID Cliente", "Nombre Operario", "Apellido Operario",
              "Cantidad de Chips(KG)", "Cantidad de Troncos", "Tipo Madera", "Tiempo de produccion"])
    wb.save("Historial-Producciones.xlsx")


# Agregamos pedido a la tabla de pedidos
def agregarPedidos(solicitud, tama):
    wb = load_workbook("Pedidos-Clientes.xlsx")
    ws = wb.worksheets[0]
    ws_tables = 0
    for i in wb.worksheets:
        ws_tables += 1
    if ws_tables == 1:
        ws.append([solicitud.fecha, solicitud.idCli, solicitud.nombre,
                  solicitud.apellido, (solicitud.cantChips*1000), tama, solicitud.tipoMadera])
    else:
        ws = wb.create_sheet("Pedidos-Clientes")
        ws.append([solicitud.fecha, solicitud.idCli, solicitud.nombre,
                  solicitud.apellido, (solicitud.cantChips*1000), tama, solicitud.tipoMadera])
    wb.save("Pedidos-Clientes.xlsx")


# Agregamos operacion a la tabla de operaciones
def agregarProducciones(solicitud, operario1, totalTroncos, tiempo):
    wb = load_workbook("Historial-Producciones.xlsx")
    ws = wb.worksheets[0]
    ws_tables = 0
    for i in wb.worksheets:
        ws_tables += 1
    if ws_tables == 1:
        ws.append([solicitud.fecha, solicitud.idCli, operario1.nombre,
                  operario1.apellido, (solicitud.cantChips*1000), round(totalTroncos), solicitud.tipoMadera, tiempo])
    else:
        ws = wb.create_sheet("Historial-Producciones")
        ws.append([solicitud.fecha, solicitud.idCli, operario1.nombre,
                  operario1.apellido, (solicitud.cantChips*1000), round(totalTroncos), solicitud.tipoMadera, tiempo])
    wb.save("Historial-Producciones.xlsx")


# Devuelve el numero de troncos a utilizar de acuerdo al tipo de madera elegido
def devolverCantidad(solicitud):

    if((solicitud.tipoMadera == 'pino') or (solicitud.tipoMadera == 'Pino')):
        cantTroncos = (solicitud.cantChips*1000)/pino.masa
    if((solicitud.tipoMadera == 'algarrobo') or (solicitud.tipoMadera == 'Algarrobo')):
        cantTroncos = (solicitud.cantChips*1000)/algarrobo.masa
    if((solicitud.tipoMadera == 'eucalipto') or (solicitud.tipoMadera == 'Eucalipto')):
        cantTroncos = (solicitud.cantChips*1000)/eucalipto.masa
    if((solicitud.tipoMadera == 'urunday') or (solicitud.tipoMadera == 'Urunday')):
        cantTroncos = (solicitud.cantChips*1000)/Urunday.masa
    if((solicitud.tipoMadera == 'quebracho') or (solicitud.tipoMadera == 'Quebracho')):
        cantTroncos = (solicitud.cantChips*1000)/quebracho.masa

    return cantTroncos


if "Pedidos-Clientes.xlsx" not in os.listdir():
    crear_excel_pedidos()
    # Si no existe el archivo excel se cre uno

if "Historial-Producciones.xlsx" not in os.listdir():
    crear_excel_prodcciones()

print("Bienvenido a SOFTCHIPS. Para empezar Ingrese sus datos\n")
datos = ingresarCredenciales()
print("Presione una tecla para empezar \n")
msvcrt.getch()
os.system("cls")

# Funcion menu donde se muestra las distintas opciones del programa


def menu():

    print("INGRESE UNA OPCION \n")
    print('1. Iniciar Nueva Produccion\n')
    print('2. Mostrar Historial de producciones\n')
    print('3. Mostrar Historial de pedidos de los clientes\n')
    print('4. Instrucciones \n')
    print('4. Salir\n')
    opcion = int(input())

    if opcion == 1:

        solicitud = solicitudCliente.cargarSolicitud()
        print("Ingrese tamaño del chip")
        tama = float(input())
        print("\nLa solicitud cargada es: \n")
        solicitud.mostrarSolicitud()
        print('Tamaño Chip: ', tama, 'cm')

        msvcrt.getch()
        os.system("cls")
        tiempo = random.randint(3, 8)

        print("Produciendo...")
        time.sleep(5)
        os.system("cls")
        print("El tiempo que tardo en producir los chips es de: ", tiempo, 'Hs \n')
        totalTroncos = devolverCantidad(solicitud)

        print("La cantidad de troncos que se van a usar para realizar la operacion es:  ",
              round(totalTroncos))
        print('\n')
        print(
            "Presione una tecla para ver la posicion en la que se debe ubicar el camion\n")
        msvcrt.getch()
        os.system("cls")
        PosCamion(tama)
        agregarPedidos(solicitud, tama)

        agregarProducciones(solicitud, datos, totalTroncos, tiempo)

        print('Presione una tecla para continuar')
        msvcrt.getch()
        os.system("cls")

        menu()

    if opcion == 2:
        libro = pd.read_excel('Historial-Producciones.xlsx')
        print(libro[["Fecha", "ID Cliente", "Nombre Operario", "Apellido Operario",
              "Cantidad de Chips(KG)", "Cantidad de Troncos", "Tipo Madera", "Tiempo de produccion"]])

        print('Presione una tecla para volver al menu')
        msvcrt.getch()
        os.system("cls")
        menu()

    if opcion == 3:

        libro = pd.read_excel('Pedidos-Clientes.xlsx')
        print(libro[['Fecha', 'ID Cliente', 'Nombre Cliente', 'Apellido Cliente',
              'Cantidad de Chips(KG)', 'Tamaño Chips', 'Tipo Madera']])

        print('Presione una tecla para volver al menu')
        msvcrt.getch()
        os.system("cls")
        menu()

    if opcion == 4:
        print('Paso 1: Para iniciar una nueva produccion ingrese la opcion 1. \n')
        print("Paso 2: Complete cada uno de los campos solicitados con la informacion del cliente. \n")
        print('Paso 3: A continuacion podra observar la cantidad de troncos que tendra que se utilizaron para dicha produccion y la posicion en la que se tendra que ubicar el camion. \n')
        print("Para observar el historial de producciones ingrese la opcion 2. \n")
        print("Para observar el historial de pedidos de los clientes ingrese la opcion 3.\n")
        print("Para Salir del programa ingrese la opcion 5.\n")
        print('Presione una tecla para volver al menu \n')
        msvcrt.getch()
        os.system("cls")
        menu()



menu()
